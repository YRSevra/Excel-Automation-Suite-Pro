from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class ExcelFormatter:

    def format_excel(self, file_path):

        wb = load_workbook(file_path)
        ws = wb.active

        # Header Style
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        thin = Side(border_style="thin", color="000000")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

        # Data Formatting
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(
                    left=thin,
                    right=thin,
                    top=thin,
                    bottom=thin
                )

        # Auto Width
        for column in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
            ws.column_dimensions[column[0].column_letter].width = length + 5

        # Freeze Header
        ws.freeze_panes = "A2"

        # Filter
        ws.auto_filter.ref = ws.dimensions

        wb.save(file_path)